from venv import logger
from rest_framework import status
from django.contrib.auth import authenticate
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from finalone.models import User, Role  # Adjust the import based on your app name
from finalone.models import User  # Adjust based on your app name
from rest_framework_simplejwt.tokens import RefreshToken # type: ignore
from rest_framework.permissions import IsAuthenticated
from finalone.permissions import IsAdminUser 
import openpyxl
from django.http import HttpResponse
from finalone.models import User
from rest_framework.parsers import MultiPartParser, FormParser





class SignupAPIView(APIView):
    
    def post(self, request, *args, **kwargs):
        print("calling.....ss")
        name = request.data.get('name')
        password = request.data.get('password')
        confirm_password = request.data.get('confirm_password')
        role_name = request.data.get('role')

        # Validate input
        if not name or not password or not confirm_password or not role_name:
            return Response({"error": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)
        
        if password != confirm_password:
            return Response({"error": "Passwords do not match."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the role exists
        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return Response({"error": "Role not found."}, status=status.HTTP_400_BAD_REQUEST)

        # Create the user
        try:
            user = User(
                name=name,
                role=role,
                password=make_password(password)  # Hash the password
            )
            user.save()
            return Response({"message": "User created successfully!"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



# class SignInAPIView(APIView):
#     def post(self, request, *args, **kwargs):
#         name = request.data.get('name')
#         password = request.data.get('password')

#         # Validate input
#         if not name or not password:
#             return Response({"error": "Both name and password are required."}, status=status.HTTP_400_BAD_REQUEST)
#         user = authenticate(name=name, password=password)
          
#         # Check if the user exists
#         try:
#             user = User.objects.get(name=name)
#         except User.DoesNotExist:
#             return Response({"error": "Invalid name or password."}, status=status.HTTP_400_BAD_REQUEST)

#         # Check password
#         if check_password(password, user.password):
#             refresh = RefreshToken.for_user(user)
#             return Response({
#                 'message': "Login successful!",
#                 'refresh': str(refresh),
#                 'access': str(refresh.access_token),
#             }, status=status.HTTP_200_OK)
#         else:
#             return Response({"error": "Invalid name or password."}, status=status.HTTP_400_BAD_REQUEST)

class SignInAPIView(APIView):
    def post(self, request, *args, **kwargs):
        # print("ddddddddddddd")
        # return Response("success")
        name = request.data.get('name')
        password = request.data.get('password')

        # Validate input
        if not name or not password:
            return Response({"error": "Both name and password are required."}, status=status.HTTP_400_BAD_REQUEST)
        print(request)
        # Authenticate the user
        user = authenticate(request,name=name, password=password)
        print(user,"dddddddddddddddd")
        if user is not None:
            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_200_OK)
        else:
            return Response({"error": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)



class UserCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]  
    def post(self, request, *args, **kwargs):
        name = request.data.get('name')
        password = request.data.get('password')
        role_name = request.data.get('role')

        if not name or not password or not role_name:
            return Response({"error": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return Response({"error": "Role not found."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User(name=name, role=role, password=make_password(password))
            user.save()
            return Response({"message": "User created successfully!"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UserDetailAPIView(APIView):
    def get(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            return Response({"name": user.name, "role": user.role.name}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

class UserUpdateAPIView(APIView):
    def put(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        name = request.data.get('name', user.name)
        role_name = request.data.get('role', None)

        if role_name:
            try:
                role = Role.objects.get(name=role_name)
            except Role.DoesNotExist:
                return Response({"error": "Role not found."}, status=status.HTTP_400_BAD_REQUEST)
            user.role = role

        user.name = name
        user.save()

        return Response({"message": "User updated successfully!"}, status=status.HTTP_200_OK)

class UserDeleteAPIView(APIView):
    def delete(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            user.delete()
            return Response({"message": "User deleted successfully!"}, status=status.HTTP_204_NO_CONTENT)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        




class UserListView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]  

    def get(self, request, *args, **kwargs):
        users = User.objects.all().values('name', 'role__name',)  
        return Response(users, status=status.HTTP_200_OK)
    


class ExportUsersToExcelAPIView(APIView):
    # permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request, *args, **kwargs):
        
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Users'

        headers = ['ID', 'Name', 'Role', 'Is Active', 'Is Staff']
        sheet.append(headers)

        
        users = User.objects.all().values('id', 'name', 'role__name', 'is_active', 'is_staff')

        for user in users:
            sheet.append([
                user['id'],
                user['name'],
                user['role__name'],
                user['is_active'],
                user['is_staff']
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'
        workbook.save(response)

        return response
    



# class ImportUsersFromExcelAPIView(APIView):
#     parser_classes = (MultiPartParser, FormParser)

#     def post(self, request, *args, **kwargs):
#         file = request.FILES.get('file')
        
#         if not file:
#             return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

#         # Open the uploaded Excel file
#         try:
#             workbook = openpyxl.load_workbook(file)
#             sheet = workbook.active

#             # Loop through the rows and extract data
#             for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is header
#                 user_id, name, role_name, is_active, is_staff = row

#                 # Validate and process the role
#                 try:
#                     role, created = Role.objects.get_or_create(name=role_name)
#                 except Role.DoesNotExist:
#                     return Response({"error": f"Role '{role_name}' not found."}, status=status.HTTP_400_BAD_REQUEST)

#                 # Create or update the user
#                 user, created = User.objects.update_or_create(
#                     id=user_id,  # If you want to update by ID, or skip this for new users
#                     defaults={
#                         'name': name,
#                         'role': role,
#                         'is_active': is_active,
#                         'is_staff': is_staff
#                     }
#                 )

#             return Response({"message": "Users imported successfully!"}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

class ImportExcelDataAPIView(APIView):
    def post(self, request, *args, **kwargs):
        # Assuming the Excel file is passed in the form data as 'file'
        excel_file = request.FILES['file']

        # Load the Excel file
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Extract headers (assuming headers are in the first row)
        headers = [cell.value for cell in sheet[1]]

        # List to store the rows as dictionaries
        data_list = []

        # Iterate through each row after the header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for the current row
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            data_list.append(row_dict)

        # Return the data as a JSON response
        return Response(data_list, status=200)
